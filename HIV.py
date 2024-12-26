import numpy as np
import pandas as pd
import openpyxl
import random
import math
from tqdm import tqdm
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from mpl_toolkits.mplot3d import Axes3D
from colorama import Fore,Style,init
import os


class Cell:
    def __init__(self,params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        #细胞参数一般为定值
        self.l1, self.l2 = 10, 0.04
        self.dT1, self.dT2 = 0.01, 0.01
        self.f1, self.f2 = 0.1, 0.1
        self.a1, self.a2 = 0.2, 0.05
        self.dL1, self.dL2 = 0.02, 0.1
        self.dI1, self.dI2 = 0.5, 0.1
        self.N1, self.N2 = 15, 10
        self.c1V, self.c2V = 3, 3
        #感染参数一般为变量b:1e-5-0.5;e:0-1
        self.b1, self.b2, self.e1RT, self.e2RT, self.e1PI, self.e2PI=params
        #用数组储存细胞过程
        self.cell_process = np.zeros(int(nt+1))
        self.dt=te/nt

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return i

    def process(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        self.cell_process[i+1]=(self.cell_process[i] + self.dnumber(T1, L1, I1, T2, L2, I2, VI, VNI, i)*self.dt)

    def R0(self,params=[0.0001,0.015,0,0,0,0]):
        b1, b2, e1RT, e2RT, e1PI, e2PI=params
        return (self.N1*(1-e1PI)*b1*(1-e1RT)*(self.a1+(1-self.f1)*self.dL1))/(self.c1V*(self.dL1+self.a1))*(self.l1/self.dT1)+(self.N2*(1-e2PI)*b2*(1-e2RT)*(self.a2+(1-self.f2)*self.dL2))/(self.c2V*(self.dL2+self.a2))*(self.l2/self.dT2)

class Cell_T1(Cell):
    def __init__(self, T10=1000, params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        super().__init__(params=params,te=te,nt=nt)
        self.cell_process[0] = T10

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return self.l1-self.dT1*T1.cell_process[i]-self.b1*(1-self.e1RT)*T1.cell_process[i]*VI.cell_process[i]

class Cell_L1(Cell):
    def __init__(self, L10=0, params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        super().__init__(params=params,te=te,nt=nt)
        self.cell_process[0] = L10

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return self.f1*self.b1*(1-self.e1RT)*T1.cell_process[i]*VI.cell_process[i]-self.a1*L1.cell_process[i]-self.dL1*L1.cell_process[i]

class Cell_I1(Cell):
    def __init__(self, I10=0, params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        super().__init__(params=params,te=te,nt=nt)
        self.cell_process[0] = I10

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return (1-self.f1)*self.b1*(1-self.e1RT)*T1.cell_process[i]*VI.cell_process[i]+self.a1*L1.cell_process[i]-self.dI1*I1.cell_process[i]

class Cell_T2(Cell):
    def __init__(self, T20=4, params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        super().__init__(params=params,te=te,nt=nt)
        self.cell_process[0] = T20

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return self.l2-self.dT2*T2.cell_process[i]-self.b2*(1-self.e2RT)*T2.cell_process[i]*VI.cell_process[i]

class Cell_L2(Cell):
    def __init__(self, L20=0, params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        super().__init__(params=params,te=te,nt=nt)
        self.cell_process[0] = L20

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return self.f2*self.b2*(1-self.e2RT)*T2.cell_process[i]*VI.cell_process[i]-self.a2*L2.cell_process[i]-self.dL2*L2.cell_process[i]

class Cell_I2(Cell):
    def __init__(self, I20=0, params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        super().__init__(params=params,te=te,nt=nt)
        self.cell_process[0] = I20

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return (1-self.f2)*self.b2*(1-self.e2RT)*T2.cell_process[i]*VI.cell_process[i]+self.a2*L2.cell_process[i]-self.dI2*I2.cell_process[i]

class Cell_VI(Cell):
    def __init__(self, VI0=0, params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        super().__init__(params=params,te=te,nt=nt)
        self.cell_process[0] = VI0

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return self.N1*(1-self.e1PI)*self.dI1*I1.cell_process[i]+self.N2*(1-self.e2PI)*self.dI2*I2.cell_process[i]-self.c1V*VI.cell_process[i]

class Cell_VNI(Cell):
    def __init__(self, VNI0=0, params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
        super().__init__(params=params,te=te,nt=nt)
        self.cell_process[0] = VNI0

    def dnumber(self, T1=None, L1=None, I1=None, T2=None, L2=None, I2=None, VI=None, VNI=None, i=0):
        return self.N1*self.e1PI*self.dI1*I1.cell_process[i]+self.N2*self.e2PI*self.dI2*I2.cell_process[i]-self.c2V*VNI.cell_process[i]

def all_cells_processes(cell0=[1000,0,0,4,0,0,0,0],params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000,order:'all' or 'all_noreport' or 'final'='all'):
    if (order!='all' and order!='all_noreport' and order!='final'):
        print('all_cells_processes命令不正确。')
    else:
        #计算参数
        T10, L10, I10, T20, L20, I20, VI0, VNI0 = cell0
        T1 = Cell_T1(T10=T10, params=params,te=te,nt=nt)
        L1 = Cell_L1(L10=L10, params=params,te=te,nt=nt)
        I1 = Cell_I1(I10=I10, params=params,te=te,nt=nt)
        T2 = Cell_T2(T20=T20, params=params,te=te,nt=nt)
        L2 = Cell_L2(L20=L20, params=params,te=te,nt=nt)
        I2 = Cell_I2(I20=I20, params=params,te=te,nt=nt)
        VI = Cell_VI(VI0=VI0, params=params,te=te,nt=nt)
        VNI = Cell_VNI(VNI0=VNI0, params=params,te=te,nt=nt)
        cells = [T1, L1, I1, T2, L2, I2, VI, VNI]
        for i in range(int(nt)):
            for item in cells:
                item.process(T1=T1, L1=L1, I1=I1, T2=T2, L2=L2, I2=I2, VI=VI, VNI=VNI, i=i)
    if order=='all':
        print('细胞终值：\n',{type(item).__name__:item.cell_process[-1] for item in cells},'\n')
        return [item.cell_process for item in cells]
    elif order=='all_noreport':
        return [item.cell_process for item in cells]
    elif order=='final':
        return [item.cell_process[-1] for item in cells]

def test():
    E0,R00=[1000,0,0,4,0,0,0,0],0.6821#b1=0.0001,b2=0.015时,R0=0.6821无病平衡点实际值
    EI,R0I=[335.4188,3.0208,13.1704,0.9297,0.0205,0.2866,33.0290,0],2.9852#b1=0.0006,b2=0.001时,R0=2.9852的染病平衡点实际值
    #0.1 测试cell.R0()
    print(f'\n0.1 R00的相对计算误差为：{(cell.R0([0.0001,0.015,0,0,0,0])-R00)/R00}；R0I的相对计算误差为：{(cell.R0([0.0006,0.001,0,0,0,0])-R0I)/R0I}；\n')
    #0.2 测试all_cells_processes()
    te,nt,dt=1000,10000,0.1
    print(f'0.2 te={te},nt={nt}时\n无病平衡点误差：{[item if E0[j]==0 else (item-E0[j])/E0[j] for j,item in enumerate(all_cells_processes(cell0=[500,1,3,1,0.001,0.001,12,0],params=[0.0001,0.015,0,0,0,0],te=te,nt=nt,order='final'))]}\n染病平衡点误差：{[item if EI[j]==0 else (item-EI[j])/EI[j] for j,item in enumerate(all_cells_processes(cell0=[500,1,3,1,0.001,0.001,12,0],params=[0.0006,0.001,0,0,0,0],te=te,nt=nt,order='final'))]}\n')
    #0.3 取点规模nt及dt对结果的影响
    print('0.3 计算误差随dt变化图。')
    te,nt=5000,np.linspace(10000,100000,91)#反应时长，取点规模
    dt=te/nt#微分步长
    dev_E0=[]
    for i,item in tqdm(enumerate(nt),desc='testing'):
        cells_final=all_cells_processes(cell0=[500,1,3,1,0.001,0.001,12,0],params=[0.0001,0.015,0,0,0,0],te=te,nt=int(item),order='final')
        dev_E0.append([item if E0[j]==0 else (item-E0[j])/E0[j] for j,item in enumerate(cells_final)])
    dev_E0=np.array(dev_E0).T
    fig, axs = plt.subplots(4, 2, constrained_layout=True, sharex='col')
    for i, ax in enumerate(axs.flat):
        ax.set_title(cells[i])
        if i%2==0:
            ax.set_ylabel('deviation')
        if i//2==3:
            ax.set_xlabel(r'$\delta_{t}$')
        ax.axvline(x=0.1, color='blue', linestyle='--',label=r'$\delta_{t}=1\times 10^{-1}$')
        ax.axhline(y=0, color='red', linestyle='--',label='zero')
        ax.plot(dt,dev_E0[i], linewidth=0.43, zorder=1)
        sc=ax.scatter(dt,dev_E0[i],s=10,c=nt,cmap=cmap2, zorder=2)
        ax.legend()
    cb = fig.colorbar(sc, ax=axs.ravel().tolist())
    cb.set_label(r'$n_{t}$')
    plt.show()

def params_R0(n=100,nature:bool=False,R0_min=0,R0_max=2,b1m=6e-4,b1d=2e-4,b2m=9e-2,b2d=3e-2,drug:bool=False,eRTm=0.25,eRTd=0,ePIm=0.25,ePId=0):
    if (nature!=True and nature!=False) or (drug!=True and drug!=False):
        print('all_cells_processes命令不正确。')
    else:
        k,R0_pass=0,False
        b1, b2, eRT, ePI, R0=[],[],[],[],[]
        while k<n:
            if nature:
                b1_random, b2_random = np.random.normal(b1m,b1d), np.random.normal(b2m,b2d)
                R0_pass=True
            else:
                b1_random, b2_random = random.uniform(1e-5,0.5), random.uniform(1e-5,0.5)
            if drug:
                eRT_random, ePI_random = np.random.normal(eRTm,eRTd), np.random.normal(ePIm,ePId)
            else:
                eRT_random, ePI_random = 0,0
            #仅根据b1,b2筛选样本点故eRT,ePI恒为0。
            R0_random=cell.R0([b1_random,b2_random,0,0,0,0])
            if (R0_min<=R0_random<=R0_max or R0_pass) and b1_random>=0 and b2_random>=0:
               k+=1
               b1.append(b1_random)
               b2.append(b2_random)
               eRT.append(eRT_random)
               ePI.append(ePI_random)
               #筛选后再计算R0
               R0.append(cell.R0([b1_random,b2_random,eRT_random,eRT_random,ePI_random,ePI_random]))
        sorted_index=np.argsort(R0)
        return [(np.array(b1))[sorted_index],(np.array(b2))[sorted_index],(np.array(eRT))[sorted_index],(np.array(ePI))[sorted_index],(np.array(R0))[sorted_index]]

def cells_processes_params(params_variable=None,cell0=[1000,0,0,4,0,0,0,0],order:'all' or 'final'='final',filename=None,sheet_name=str(0)):
    if order!='all' and order!='final':
        print('all_cells_processes命令不正确。')
    else:
        n=len(params_variable[0])
        if order=='final':
            R0_cells_final=[]
            for i in tqdm(range(n),desc=f'cell0:{cell0},cells processing'):
                R0_cells_final.append(all_cells_processes(cell0=cell0,params=[params_variable[0][i],params_variable[1][i],params_variable[2][i],params_variable[2][i],params_variable[3][i],params_variable[3][i]],order='final'))
            cells_final=np.array(R0_cells_final)
            if filename:
                data=np.concatenate((np.array(params_variable).T,cells_final),axis=1)
                try:
                    wb = openpyxl.load_workbook(filename=filename+'.xlsx')
                except FileNotFoundError:
                    wb = openpyxl.Workbook()
                    wb.remove(wb['Sheet'])
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    k=0
                else:
                    ws = wb.create_sheet(sheet_name)
                    k=1
                last_row = ws.max_row-k
                lrow=last_row
                print(f'\'{filename}.xlsx\'的booksheet\'{sheet_name}\'中原有{lrow}组数据',end='；')
                for row in data:
                    last_row += 1
                    for column, value in enumerate(row, start=1):
                        ws.cell(row=last_row, column=column, value=value)
                wb.save(filename=filename+'.xlsx')
                print(f'向\'{filename}.xlsx\'的booksheet\'{sheet_name}\'中添加了{last_row-lrow}组，现共有{last_row}组数据。')
            return cells_final.T
        elif order=='all':
            R0_cells_processes=[]
            for i in tqdm(range(n),desc=f'cell0:{cell0},cells processing'):
                R0_cells_processes.append(all_cells_processes(cell0=cell0,params=[params_variable[0][i],params_variable[1][i],params_variable[2][i],params_variable[2][i],params_variable[3][i],params_variable[3][i]],order='all_noreport'))
            return np.array(R0_cells_processes)

def get_from_xlsx(filename='cells_processes_params',sheet_name=str(0)):
    df=pd.read_excel(filename+'.xlsx',sheet_name=sheet_name,header=None)
    data=(df.values).T
    #根据R0进行排序
    sorted_index=np.argsort(data[4])
    sorted_data=data[:,sorted_index]
    print(f'从\'{filename}.xlsx\'的booksheet\'{sheet_name}\'中读取了{len(sorted_index)}组数据。')
    return sorted_data[:5],sorted_data[5:]

def generate_final_data_xlsx(filename=None,n=20,nature:bool=False,R0_min=0,R0_max=2,b1m=6e-4,b1d=2e-4,b2m=9e-2,b2d=3e-2,drug_kind:'given'or'eRT'or'ePI'or'both_drug'=None,eRTm=0.5,ePIm=0.5,n_drug=100):
    if drug_kind!='given' and drug_kind!='eRT' and drug_kind!='ePI' and drug_kind!='both_drug' and drug_kind!=None and filename:
        print('generate_final_data_xlsx命令不正确。')
    else:
        cell0=[500,1,3,1,0.001,0.001,12,0]
        if drug_kind=='given':
            params_variable=params_R0(R0_min=R0_min,R0_max=R0_max,n=n,nature=nature,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug=True,eRTm=eRTm,ePIm=ePIm)
            cells_processes_params(params_variable=params_variable,cell0=cell0,order='final',filename=filename)
        elif drug_kind=='eRT':
            for i in range(n_drug):
                params_variable=params_R0(n=n,nature=nature,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug=True,eRTm=(1/(n_drug-1))*(i),ePIm=0)
                cells_processes_params(params_variable=params_variable,cell0=cell0,order='final',filename=filename,sheet_name=str(i))
        elif drug_kind=='ePI':
            for i in range(n_drug):
                params_variable=params_R0(n=n,nature=nature,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug=True,eRTm=0,ePIm=(1/(n_drug-1))*(i))
                cells_processes_params(params_variable=params_variable,cell0=cell0,order='final',filename=filename,sheet_name=str(i))
        elif drug_kind=='both_drug':
            n_point=int(math.sqrt(n_drug))
            for i in range(n_point):
                for j in range(n_point):
                    params_variable=params_R0(n=n,nature=nature,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug=True,eRTm=(1/(n_point-1))*(i),ePIm=(1/(n_point-1))*(j))
                    cells_processes_params(params_variable=params_variable,cell0=cell0,order='final',filename=filename,sheet_name=str((i)*n_point+j))
        else:
            params_variable=params_R0(R0_min=R0_min,R0_max=R0_max,n=n,nature=nature,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug=False)
            cells_processes_params(params_variable=params_variable,cell0=cell0,order='final',filename=filename)

def healing_rate(order:'rate_params' or 'rate'='rate',params_variable=None,cells_final=None,T1_final=None,T2_final=None,tol=1e-3):
    if order=='rate_params':
        k1,k2=0,0
        healing_b1,healing_b2,healing_eRT,healing_ePI,healing_R0=[],[],[],[],[]
        healing_params=[healing_b1,healing_b2,healing_eRT,healing_ePI,healing_R0]
        nuhealing_b1,nuhealing_b2,nuhealing_eRT,nuhealing_ePI,nuhealing_R0=[],[],[],[],[]
        nuhealing_params=[nuhealing_b1,nuhealing_b2,nuhealing_eRT,nuhealing_ePI,nuhealing_R0]
        for i,R0 in enumerate(params_variable[4]):
            #T1与T2接近正常值即可认为治愈
            if (cells_final[0][i]/1e3)>1-tol and cells_final[3][i]/4>1-tol:
                k1+=1
                for j,item in enumerate(healing_params):
                    item.append(params_variable[j][i])
            else:
                k2+=1
                for j,item in enumerate(nuhealing_params):
                    item.append(params_variable[j][i])
        return k1/(k1+k2),healing_params,nuhealing_params
    elif order=='rate':
        k1,k2=0,0
        for i,item in enumerate(T1_final):
            if (item/1e3)>1-tol and (T2_final[i]/4)>1-tol:
                k1+=1
            else:
                k2+=1
        return k1/(k1+k2)

def fig_cells_processing_cell0s(cell0s=[[1000,0,0,4,0,0,0,0]],params=[0.0001,0.015,0,0,0,0],te=1000,nt=10000):
    times=np.linspace(0,te,nt+1)
    cell_processes=dict({})
    for i,item in enumerate(cell0s):
        print(f'初始状态{i+1}：cell0:{item};params:{params};R0={cell.R0(params=params)}')
        cell_processes[i]=all_cells_processes(cell0=item,params=params,te=te,nt=nt,order='all')
    fig, axs = plt.subplots(4, 2, constrained_layout=True, sharex='col')
    for i, ax in enumerate(axs.flat):
        ax.set_title(cells[i])
        if i%2==0:
            ax.set_ylabel('concentration')
        if i//2==3:
            ax.set_xlabel('t/day')
        for j,item in enumerate(cell0s):
            ax.plot(times,cell_processes[j][i],label=j+1)
        ax.legend()
    plt.show()

def fig_cells_processing_paramses(cell0=[1000,0,0,4,0,0,0,0],paramses=[[0.0001,0.015,0,0,0,0]],te=1000,nt=10000):
    times=np.linspace(0,te,nt+1)
    cell_processes=dict({})
    for i,item in enumerate(paramses):
        print(f'初始状态{i+1}：cell0:{cell0};params:{item};R0={cell.R0(params=item)}')
        cell_processes[i]=all_cells_processes(cell0=cell0,params=item,te=te,nt=nt,order='all')
    fig, axs = plt.subplots(4, 2, constrained_layout=True, sharex='col')
    for i, ax in enumerate(axs.flat):
        ax.set_title(cells[i])
        if i%2==0:
            ax.set_ylabel('concentration')
        if i//2==3:
            ax.set_xlabel('t/day')
        for j,item in enumerate(paramses):
            ax.plot(times,cell_processes[j][i],label=j+1)
        ax.legend()
    plt.show()

def fig_final_R0(params_variables=None,cells_finals=None):
    fig, axs = plt.subplots(4, 2, constrained_layout=True,sharex='col')
    if len(cells_finals)==1:
        for i, ax in enumerate(axs.flat):
            ax.set_title(cells[i])
            ax.axvline(x=1, color='red', linestyle='--',label=r'$R_{0}=1$')
            ax.plot(params_variables[0][4],cells_finals[0][i], linewidth=0.43, zorder=1)
            sc=ax.scatter(params_variables[0][4],cells_finals[0][i],s=10,c=params_variables[0][4],cmap=cmap2, zorder=2)
            if i%2==0:
                ax.set_ylabel('concentration')
            if i//2==3:
                ax.set_xlabel(r'$R_{0}$')
        cb = fig.colorbar(sc, ax=axs.ravel().tolist())
        cb.set_label(r'$R_{0}$')
    else:
        for i, ax in enumerate(axs.flat):
            ax.set_title(cells[i])
            ax.axvline(x=1, color='red', linestyle='--',label=r'$R_{0}=1$')
            if i%2==0:
                ax.set_ylabel('concentration')
            if i//2==3:
                ax.set_xlabel(r'$R_{0}$')
            for j,item in enumerate(cell0s):
                ax.plot(params_variables[j][4],cells_finals[j][i],label=j+1)
                ax.legend()
    plt.show()

def fig_R0_b1_b2(params_variable=None):
    fig,ax=plt.subplots(subplot_kw={'projection': '3d'})#R0-b1,b2三维
    sc=ax.scatter(params_variable[0], params_variable[1], params_variable[4], c=params_variable[4], cmap=cmap2)
    ax.set_xlabel(r'$\beta_{1}$')
    ax.set_ylabel(r'$\beta_{2}$')
    ax.set_zlabel(r'$R_{0}$')
    fig.colorbar(sc, label=r'$R_{0}$')
    fig,ax=plt.subplots()#R0-b1,b2二维
    sc=ax.scatter(params_variable[0], params_variable[1], c=params_variable[4], cmap=cmap2)
    ax.set_xlabel(r'$\beta_{1}$')
    ax.set_ylabel(r'$\beta_{2}$')
    fig.colorbar(sc, label=r'$R_{0}$')
    plt.show()

def fig_final_b1_b2(params_variable=None,cells_final=None):
    fig, axs = plt.subplots(2, 4, subplot_kw={'projection': '3d'}, constrained_layout=True)#cell状态与b1,b2关系
    for i,ax in enumerate(axs.flat):
        sc=ax.scatter(params_variable[0], params_variable[1], cells_final[i], c=params_variable[4], cmap=cmap2)
        ax.set_title(cells[i])
        ax.set_xlabel(r'$\beta_{1}$')
        ax.set_ylabel(r'$\beta_{2}$')
        if i%4==3:
            ax.set_zlabel('concentration')
    cb = fig.colorbar(sc, ax=axs.ravel().tolist())
    cb.set_label(r'$R_{0}$')
    plt.show()

def fig_drug_effect(b1=0.0006,b2=0.015):
    n=int(1e5)
    eRT = np.random.uniform(0, 1, size=n)
    ePI=np.linspace(0, 1, n)
    R0=np.zeros(n)
    R1_x,R1_y=[],[]
    for i in range(n):
        R0[i]=cell.R0([b1,b2,eRT[i],eRT[i],ePI[i],ePI[i]])
        if abs(R0[i]-1)<=8e-3:
            R1_x.append(eRT[i])
            R1_y.append(ePI[i])
    fig, ax = plt.subplots()
    sc1 = ax.scatter(eRT, ePI, c=R0, cmap=cmap1, s=2)
    sc2 = ax.scatter(R1_x, R1_y, color='darkred',s=0.8,label=r'$R_{0}=1$')
    fig.colorbar(sc1, label=r'$R_{0}$')
    ax.set_xlabel(r"$\varepsilon _{RT}$")
    ax.set_ylabel(r"$\varepsilon _{PI}$")
    plt.legend()
    plt.show()

def fig_compare_healing_b1_b2(params_variable_nodrug=None,cells_final_nodrug=None,params_variable_drug=None,cells_final_drug=None,tol=1e-2):
    healing_rate_nodrug,healing_params_nodrug,nuhealing_params_nodrug=healing_rate(order='rate_params',params_variable=params_variable_nodrug,cells_final=cells_final_nodrug,tol=tol)
    healing_rate_drug,healing_params_drug,nuhealing_params_drug=healing_rate(order='rate_params',params_variable=params_variable_drug,cells_final=cells_final_drug,tol=tol)
    fig,ax=plt.subplots(subplot_kw={'projection': '3d'})#R0-b1,b2三维
    #R0=1参照面
    xx, yy = np.meshgrid(np.linspace(0, max(nuhealing_params_drug[0]), 2), np.linspace(0, max(nuhealing_params_drug[1]), 2))
    zz = 0*xx+1
    ax.plot_surface(xx, yy, zz, color='r', alpha=0.25,label=r'$R_{0}=1$')
    ax.scatter(healing_params_nodrug[0], healing_params_nodrug[1], healing_params_nodrug[4],marker='^', color='blue',label='no drug healing point')
    ax.scatter(nuhealing_params_nodrug[0], nuhealing_params_nodrug[1], nuhealing_params_nodrug[4],marker='^', color='red',label='no drug point')
    ax.scatter(healing_params_drug[0], healing_params_drug[1], healing_params_drug[4],marker='o', color='blue',label='having drug healing point')
    ax.scatter(nuhealing_params_drug[0], nuhealing_params_drug[1], nuhealing_params_drug[4],marker='o', color='red',label='having drug point')
    ax.set_title(f'No Drug Healing Rate:{healing_rate_nodrug*100:.2f}%; Having Drug Healing Rate:{healing_rate_drug*100:.2f}%')
    ax.set_xlabel(r'$\beta_{1}$')
    ax.set_ylabel(r'$\beta_{2}$')
    ax.set_zlabel(r'$R_{0}$')
    ax.legend()
    fig,axs=plt.subplots(1,2,constrained_layout=True, sharex='all', sharey='all')#R0-b1,b2二维
    axs[0].scatter(healing_params_nodrug[0], healing_params_nodrug[1],marker='^', color='blue',label=' no drug healing point')
    axs[0].scatter(nuhealing_params_nodrug[0], nuhealing_params_nodrug[1],marker='^', color='red',label=' no drug point')
    axs[0].set_title(f'No Drug Healing Rate:{healing_rate_nodrug*100:.2f}%')
    axs[0].set_xlabel(r'$\beta_{1}$')
    axs[0].set_ylabel(r'$\beta_{2}$')
    axs[0].legend()
    axs[1].scatter(healing_params_drug[0], healing_params_drug[1],marker='o', color='blue',label='having drug healing point')
    axs[1].scatter(nuhealing_params_drug[0], nuhealing_params_drug[1],marker='o', color='red',label='having drug point')
    axs[1].set_title(f'Having Drug Healing Rate:{healing_rate_drug*100:.2f}%')
    axs[1].set_xlabel(r'$\beta_{1}$')
    axs[1].legend()
    plt.show()

def fig_rate_drug_nature():
    #单药
    edrug1=np.linspace(0,1,n_drug1)
    #自然条件且仅有逆转录酶抑制剂
    T1_final_RT,T2_final_RT=[],[]
    healing_rates_RT=np.zeros(n_drug1)
    for i in tqdm(range(n_drug1),desc='geting data from \'自然条件且仅有逆转录酶抑制剂\''):
        data = pd.read_excel('自然条件且仅有逆转录酶抑制剂'+'.xlsx', sheet_name=str(i),header=None)
        T1_final_RT.append((data.iloc[:, 5]).to_numpy())
        T2_final_RT.append((data.iloc[:, 8]).to_numpy())
    for i,item in tqdm(enumerate(T1_final_RT),desc='calculating healing rate just eRT'):
        healing_rates_RT[i]=healing_rate(order='rate',T1_final=item,T2_final=T2_final_RT[i])
    #自然条件且仅有蛋白酶抑制剂
    T1_final_PI,T2_final_PI=[],[]
    healing_rates_PI=np.zeros(n_drug1)
    for i in tqdm(range(n_drug1),desc='geting data from \'自然条件且仅有蛋白酶抑制剂\''):
        data = pd.read_excel('自然条件且仅有蛋白酶抑制剂'+'.xlsx', sheet_name=str(i),header=None)
        T1_final_PI.append((data.iloc[:, 5]).to_numpy())
        T2_final_PI.append((data.iloc[:, 8]).to_numpy())
    for i,item in tqdm(enumerate(T1_final_PI),desc='calculating healing rate just ePI'):
        healing_rates_PI[i]=healing_rate(order='rate',T1_final=item,T2_final=T2_final_PI[i])
    fig, axs = plt.subplots(1,2, constrained_layout=True, sharex='all',sharey='all')
    sc=axs[0].scatter(edrug1, healing_rates_RT,marker='^',c=healing_rates_RT, cmap=cmap3, zorder=2)
    axs[0].plot(edrug1, healing_rates_RT, zorder=1)
    axs[0].set_title(r'$Healing Rate-\varepsilon _{RT}$')
    axs[0].set_xlabel(r'$\varepsilon _{RT}$')
    axs[0].set_ylabel('Healing Rate/%')
    sc=axs[1].scatter(edrug1, healing_rates_PI,marker='o',c=healing_rates_PI, cmap=cmap3, zorder=2)
    axs[1].plot(edrug1, healing_rates_PI, zorder=1)
    axs[1].set_title(r'$Healing Rate-\varepsilon _{PI}$')
    axs[1].set_xlabel(r'$\varepsilon _{PI}$')
    cb = fig.colorbar(sc, ax=axs.ravel().tolist())
    cb.set_label('Healing Rate/%')
    #双药
    n_point=math.sqrt(n_drug2)
    edrug2=[[(1/(n_point-1))*(i//n_point) for i in range(n_drug2)],[(1/(n_point-1))*(i%n_point) for i in range(n_drug2)]]
    #自然条件且仅有逆转录酶抑制剂
    T1_final_drug2,T2_final_drug2=[],[]
    healing_rates_drug2=np.zeros(n_drug2)
    for i in tqdm(range(n_drug2),desc='geting data from \'自然条件且用两种药\''):
        data = pd.read_excel('自然条件且用两种药'+'.xlsx', sheet_name=str(i),header=None)
        T1_final_drug2.append((data.iloc[:, 5]).to_numpy())
        T2_final_drug2.append((data.iloc[:, 8]).to_numpy())
    for i,item in tqdm(enumerate(T1_final_drug2),desc='calculating healing rate two drug'):
        healing_rates_drug2[i]=healing_rate(order='rate',T1_final=item,T2_final=T2_final_drug2[i])
    fig,ax=plt.subplots(subplot_kw={'projection': '3d'})#healing_rate_eRT_ePI三维
    sc=ax.scatter(edrug2[0], edrug2[1], healing_rates_drug2, c=healing_rates_drug2, cmap=cmap3)
    ax.set_xlabel(r'$\varepsilon _{RT}$')
    ax.set_ylabel(r'$\varepsilon _{PI}$')
    ax.set_zlabel('Healing Rate/%')
    ax.set_title(r'$Healing Rate-\varepsilon _{RT},\varepsilon _{PI}$')
    fig.colorbar(sc, label='Healing Rate/%')
    #拟合曲面
    xx=np.array(edrug2[0]).reshape(int(n_point),int(n_point))
    yy=np.array(edrug2[1]).reshape(int(n_point),int(n_point))
    zz=np.array(healing_rates_drug2).reshape(int(n_point),int(n_point))
    ax.plot_surface(xx,yy,zz, color='blue',alpha=0.1)
    fig,ax=plt.subplots()#healing_rate_eRT_ePI二维
    sc=ax.scatter(edrug2[0], edrug2[1], c=healing_rates_drug2, cmap=cmap3,s=500)
    ax.set_xlabel(r'$\varepsilon _{RT}$')
    ax.set_ylabel(r'$\varepsilon _{PI}$')
    ax.set_title(r'$Healing Rate-\varepsilon _{RT},\varepsilon _{PI}$')
    fig.colorbar(sc, label='Healing Rate/%')
    plt.show()

if __name__ == '__main__':
    #细胞种类
    cells = ['T1', 'L1', 'I1', 'T2', 'L2', 'I2', 'VI', 'VNI']
    #各细胞初值T10, L10, I10, T20, L20, I20, VI0, VNI0
    cell0s=np.array([[900,2,7,2,0.15,0.5,18,0],[700,1.5,5,1.5,0.0005,0.0005,15,0],[500,1,3,1,0.001,0.001,12,0],[300,0.5,1.5,0.5,0.1,0.3,8,0]])
    #感染参数初值b1, b2, e1RT, e2RT, e1PI, e2PI
    params=[0.0001,0.015,0,0,0,0]#无药时参数
    params_drug=[[0.0006,0.001,0.13,0.13,0.23,0.23],[0.0006,0.001,0.23,0.23,0.33,0.33],[0.0006,0.001,0.33,0.33,0.43,0.43],[0.0006,0.001,0.43,0.43,0.53,0.53]]#给药时参数
    #细胞过程时间分布
    te,nt=1000,10000#终止时间和总取点数
    cell=Cell()
    #色条
    cmap1 = LinearSegmentedColormap.from_list('camp1', ["blue", "lightgreen", "yellow", "orange"])
    cmap2 = LinearSegmentedColormap.from_list('camp2', ["blue", "yellow", "red"])
    cmap3 = LinearSegmentedColormap.from_list('camp3', [ "red",'yellow',"blue"])
    #生成数据所用参数
    b1m,b1d,b2m,b2d=6e-4,2e-4,9e-2,3e-2#b1,b2的正态分布均值和方差
    n_drug1,n_drug2=int(40+1),int((10+1)**2)#药效变化种类

    #检查数据文件参数
    init(autoreset=True)
    prompt1=f'（{Fore.GREEN}所有数据文件均齐全。{Style.RESET_ALL}）'
    prompt2=f'（{Fore.LIGHTRED_EX}！！！缺少数据文件，请先生成数据文件后再绘图！！！{Style.RESET_ALL}）'
    prompt3=f'（{Fore.LIGHTRED_EX}缺少数据文件时请先生成数据{Style.RESET_ALL}）'
    files = ['data_final_R0.xlsx', '理想条件且无药.xlsx', '理想条件且有药(eRTm=0.25,ePIm=0.25).xlsx', '自然条件且仅有蛋白酶抑制剂.xlsx', '自然条件且仅有逆转录酶抑制剂.xlsx', '自然条件且无药.xlsx', '自然条件且用两种药.xlsx', '自然条件且有药(eRTm=0.5,ePIm=0.5).xlsx']

    k=True
    while k:
        #检查数据文件是否齐全
        file_all=True
        for file in files:
            if not os.path.exists(file):
                file_all=False
                print(f"{file}缺失！")
        if file_all:
            print(prompt1+'\n')
        else:
            print(prompt2+'\n')

        a=input('0.测试，请输入：0\n1.绘制各细胞及病毒浓度变化图，请输入：1\n2.绘制感染终态与R0关系图，请输入：2\n3.绘制R0与b1, b2关系图，请输入：3\n4.绘制感染终态与b1, b2关系图，请输入：4\n5.绘制R0与eRT, ePI关系图，请输入：5\n6.绘制有无药物时治愈率对比图，请输入：6\n7.绘制自然条件下治愈率与药效关系图，请输入：7\n8.生成绘图用数据'+prompt3+'，请输入：8\n*结束程序，请输入：e\n请输入：')

        if  a=='0':
            test()

        elif a=='1':
        #1.不给药时感染过程
            print('\n1.1 params为b1=0.0001,b2=0.015,e1RT=0,e2RT=0,e1PI=0,e2PI=0时的感染过程图（1000天）。\n')
            fig_cells_processing_cell0s(cell0s=cell0s,params=[0.0001,0.015,0,0,0,0],te=800,nt=16000)#params=[0.0001,0.015,0,0,0,0]
            print('\n1.2 params为b1=0.0006,b2=0.001,e1RT=0,e2RT=0,e1PI=0,e2PI=0时的感染过程图。\n')
            fig_cells_processing_cell0s(cell0s=cell0s,params=[0.0006,0.001,0,0,0,0],te=500,nt=10000)#params=[0.0006,0.001,0,0,0,0]
        #给药时感染过程
            print('\n1.3 params为T1=500, L1=1, I1=3, T2=1, L2=0.001, I2=0.001, VI=12, VNI=0时且给药的感染过程图。\n')
            fig_cells_processing_paramses(cell0=[500,1,3,1,0.001,0.001,12,0],paramses=params_drug,te=1500,nt=30000)

        elif a=='2':
        #2.R0决定感染状态（相同params，不同cell0）
           #params_variables,cells_finals={},{}#实时生成数据
           #params_variable=params_R0(n=100,nature=False,R0_min=0,R0_max=2,drug=False)
           #for i,item in enumerate(cell0s):
                #params_variables[i],cells_finals[i]=cells_processes_params(params_variable=params_variable,cell0=item,order='final',filename='data_final_R0',sheet_name=str(i))
        #读取已有数据
            params_variables,cells_finals={},{}
            for i,item in enumerate(cell0s):
                params_variables[i],cells_finals[i]=get_from_xlsx(filename='data_final_R0',sheet_name=str(i))
            print('\n2.1 不同params（b1,b2,R0）且无药下四种细胞初值相应的终态图。\n')
            fig_final_R0(params_variables=params_variables,cells_finals=cells_finals)
            print('\n2.2 不同params（b1,b2,R0）且无药下，细胞初值为T1=500, L1=1, I1=3, T2=1, L2=0.001, I2=0.001, VI=12, VNI=0时相应的终态图。\n')
            fig_final_R0(params_variables=params_variables,cells_finals=[cells_finals[2]])

        elif a=='3':
        #3.无药时R0由b1,b2决定
        #理想条件下
            params_variable=params_R0(n=5000,nature=False,R0_min=0,R0_max=2,drug=False)
            print('\n3.1 理想条件且无药下R0随b1,b2变化分布图。\n')
            fig_R0_b1_b2(params_variable)
        #自然条件下
            params_variable=params_R0(n=5000,nature=True,b1m=6e-4,b1d=2e-4,b2m=9e-2,b2d=3e-2,drug=False)
            print('\n3.2 自然条件且无药下R0随b1,b2变化分布图。\n')
            fig_R0_b1_b2(params_variable)

        elif a=='4':
        #4.同上无药时感染状态由b1,b2决定
        #理想条件
        #可以实时计算数据
            #params_variable=params_R0(R0_min=0,R0_max=2,n=50,nature=False,drug=False)
            #cells_final=cells_processes_params(params_variable=params_variable,cell0=[500,1,3,1,0.001,0.001,12,0],order='final')
        #也可以从已有数据中读取数据
            params_variable,cells_final=get_from_xlsx(filename='理想条件且无药')
            print('\n4.1 理想条件下细胞终态随b1,b2分布图。\n')
            fig_final_b1_b2(params_variable=params_variable,cells_final=cells_final)
        #自然条件
        #可以实时计算数据
            #params_variable=params_R0(n=100,nature=True,b1m=6e-4,b1d=2e-4,b2m=9e-2,b2d=3e-2,drug=False)
            #cells_final=cells_processes_params(params_variable=params_variable,cell0=[500,1,3,1,0.001,0.001,12,0],order='final')
        #也可以从已有数据中读取数据
            params_variable,cells_final=get_from_xlsx(filename='自然条件且无药')
            print('\n4.2 自然条件下细胞终态随b1,b2分布图。\n')
            fig_final_b1_b2(params_variable=params_variable,cells_final=cells_final)

        elif a=='5':
        #5.药效表现为在b1,b2(b1=0.0006,b2=0.015)一定时减小R0
            print('\n5 在b1=0.0006,b2=0.015的情况下R0随eRT,ePI变化图。\n')
            fig_drug_effect(b1=0.0006,b2=0.015)

        elif a=='6':
        #6.有无药治愈率对比
        #6.1理想条件下
        #可以实时计算数据
            #params_variable_nodrug=params_R0(R0_min=0,R0_max=2,n=20,nature=False,drug=False)#无药时治愈率
            #cells_final_nodrug=cells_processes_params(params_variable=params_variable_nodrug,cell0=[500,1,3,1,0.001,0.001,12,0],order='final')
            #params_variable_drug=params_R0(R0_min=0,R0_max=2,n=20,nature=False,drug=True)#有药时治愈率
            #cells_final_drug=cells_processes_params(params_variable=params_variable_drug,cell0=[500,1,3,1,0.001,0.001,12,0],order='final')
        #从已有数据中读取数据
            params_variable_nodrug,cells_final_nodrug=get_from_xlsx(filename='理想条件且无药')
            params_variable_drug,cells_final_drug=get_from_xlsx(filename='理想条件且有药(eRTm=0.25,ePIm=0.25)')
            print('\n6.1 理想条件下有药与无药治愈率对比图。\n')
            fig_compare_healing_b1_b2(params_variable_nodrug=params_variable_nodrug,cells_final_nodrug=cells_final_nodrug,params_variable_drug=params_variable_drug,cells_final_drug=cells_final_drug,tol=1e-3)
        #6.2自然条件下
        #可以实时计算数据
            #params_variable_nodrug=params_R0(n=50,nature=True,b1m=6e-4,b1d=2e-4,b2m=9e-2,b2d=3e-2,drug=False) #无药时治愈率
            #cells_final_nodrug=cells_processes_params(params_variable=params_variable_nodrug,cell0=[500,1,3,1,0.001,0.001,12,0],order='final')
            #params_variable_drug=params_R0(n=50,nature=True,b1m=6e-4,b1d=2e-4,b2m=9e-2,b2d=3e-2,drug=True,eRTm=0.5,ePIm=0.5)#有药时治愈率
            #cells_final_drug=cells_processes_params(params_variable=params_variable_drug,cell0=[500,1,3,1,0.001,0.001,12,0],order='final')
        #从已有数据中读取数据
            params_variable_nodrug,cells_final_nodrug=get_from_xlsx(filename='自然条件且无药')
            params_variable_drug,cells_final_drug=get_from_xlsx(filename='自然条件且有药(eRTm=0.5,ePIm=0.5)')
            print('\n6.2 自然条件下有药与无药治愈率对比图。\n')
            fig_compare_healing_b1_b2(params_variable_nodrug=params_variable_nodrug,cells_final_nodrug=cells_final_nodrug,params_variable_drug=params_variable_drug,cells_final_drug=cells_final_drug,tol=1e-3)

        elif a=='7':
        #7.自然条件下治愈率:单种药与两种药共用药效
            print('\n7 治愈率与药效关系图。\n')
            fig_rate_drug_nature()

        elif a=='8':
        #0.生成数据('.xlsx'文件中A:b1;B:b2;C:eRT;D:ePI;E:R0;F:T1;G:L1;H:I1;I:T2;J:L2;K:I2;L:VI;M:VNI)
        #0.1定params而变cell0所得数据，用于绘制细胞终值与R0关系图
            params_variable=params_R0(n=100,nature=False,R0_min=0,R0_max=2,drug=False)
            for i,item in enumerate(cell0s):
                cells_processes_params(params_variable=params_variable,cell0=item,order='final',filename='data_final_R0',sheet_name=str(i))
        #0.2'理想条件无药'的终态数据（理想条件：假设b1,b2在（1e-5-0.5）范围均匀分布；)
            generate_final_data_xlsx(filename='理想条件且无药',n=100,nature=False,R0_min=0,R0_max=2,drug_kind=None)
        #0.3'理想条件且有药(eRTm=0.25,ePIm=0.25)'的终态数据
            generate_final_data_xlsx(filename='理想条件且有药(eRTm=0.25,ePIm=0.25)',n=100,nature=False,R0_min=0,R0_max=2,drug_kind='given',eRTm=0.25,ePIm=0.25)
        #(自然条件b1,b2分别服从正态分布（6e-4,2e-4），（9e-2,3e-2）)
        #0.4'自然条件无药'的终态数据（理想条件：假设b1,b2在（1e-5-0.5）范围均匀分布；)
            generate_final_data_xlsx(filename='自然条件且无药',n=100,nature=True,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug_kind=None)
        #0.5'自然条件且有药(eRTm=0.25,ePIm=0.25)'的终态数据
            generate_final_data_xlsx(filename='自然条件且有药(eRTm=0.5,ePIm=0.5)',n=100,nature=True,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug_kind='given',eRTm=0.5,ePIm=0.5)
        #0.6'自然条件且仅有逆转录酶抑制剂'药效变化步长为2.5%
            generate_final_data_xlsx(filename='自然条件且仅有逆转录酶抑制剂',n=50,nature=True,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug_kind='eRT',n_drug=n_drug1)
        #0.7'自然条件且仅有蛋白酶抑制剂'药效变化步长为2.5%
            generate_final_data_xlsx(filename='自然条件且仅有蛋白酶抑制剂',n=50,nature=True,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug_kind='ePI',n_drug=n_drug1)
        #0.8'自然条件且用两种药'药效变化步长为10%
            generate_final_data_xlsx(filename='自然条件且用两种药',n=50,nature=True,b1m=b1m,b1d=b1d,b2m=b2m,b2d=b2d,drug_kind='both_drug',n_drug=n_drug2)

            print('\n一轮数据生成完成。\n得到\'.xlsx\'文件中各列储存数据为：A:b1； B:b2； C:eRT； D:ePI； E:R0； F:T1； G:L1； H:I1； I:T2； J:L2； K:I2； L:VI； M:VNI；\n')

        elif a=='e':
            print('程序运行结束。\n')
            k=False

        else:
            print('输入指令有误，请重新输入。\n')